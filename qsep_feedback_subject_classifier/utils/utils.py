import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet



def xlsx_to_dataframe(file_path: str) -> pd.DataFrame:
    """
    Read an XLSX file and return a DataFrame.

    Args:
        file_path (str): The path to the XLSX file.

    Returns:
        pd.DataFrame: The DataFrame containing the data from the XLSX file.
    """
    return pd.read_excel(file_path)

def xlsx_to_dataframes_dict(file_path: str) -> dict[str, pd.DataFrame]:
    """
    Read all sheets from an XLSX file and return a dictionary of DataFrames.

    Args:
        file_path (str): The path to the XLSX file.

    Returns:
        dict[str, pd.DataFrame]: Dictionary with sheet names as keys and DataFrames as values.
    """
    excel_file = pd.ExcelFile(file_path)
    sheet_names = excel_file.sheet_names
    
    dataframes_dict = {}
    for sheet_name in sheet_names:
        sheet_dataframe = pd.read_excel(file_path, sheet_name=sheet_name)
        dataframes_dict[sheet_name] = sheet_dataframe
    
    return dataframes_dict

def dataframe_to_xlsx(dataframe: pd.DataFrame, output_file_path: str):
    """
    Write a DataFrame to an XLSX file.

    Args:
        dataframe (pd.DataFrame): The DataFrame to write.
        output_file_path (str): The path where the XLSX file will be saved.
    """
    dataframe.to_excel(output_file_path, index=False)

def dataframes_dict_to_xlsx(dataframes_dict: dict[str, pd.DataFrame], output_file_path: str):
    """
    Write multiple DataFrames to separate sheets in an XLSX file.

    Args:
        dataframes_dict (dict[str, pd.DataFrame]): Dictionary with sheet names as keys and DataFrames as values.
        output_file_path (str): The path where the XLSX file will be saved.
    """
    excel_writer = pd.ExcelWriter(output_file_path, engine='openpyxl')
    
    for sheet_name, dataframe in dataframes_dict.items():
        dataframe.to_excel(excel_writer, sheet_name=sheet_name, index=False)
    
    excel_writer.close()


def unmerge_cells_and_fill(ws: Worksheet):
   
    for merged_range in list(ws.merged_cells.ranges):
        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        value = top_left.value
        ws.unmerge_cells(str(merged_range))
        for row in ws.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                cell.value = value


def trim_rows_until_subject(ws: Worksheet, keyword="Subject"):

    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == keyword.lower():
                header_row_idx = cell.row
                break
        if header_row_idx:
            break

    if header_row_idx is None:
        raise ValueError(f"Could not find a row with '{keyword}'")

    if header_row_idx > 1:
        ws.delete_rows(1, header_row_idx - 1)

def delete_columns_with_no_header(ws: Worksheet):
   
    for col_idx in range(ws.max_column, 0, -1):
        header_cell = ws.cell(row=1, column=col_idx)
        if not header_cell.value or str(header_cell.value).strip() == "":
            ws.delete_cols(col_idx)


def normalize_excel_and_save(file_path: str, output_path: str) -> None:
   
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        unmerge_cells_and_fill(ws)
        trim_rows_until_subject(ws)
        delete_columns_with_no_header(ws)

    # Save cleaned workbook
    wb.save(output_path)
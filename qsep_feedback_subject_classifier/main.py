import phrase_classifier
import pandas as pd
import pyperclip
from pathlib import Path
from qsep_feedback_subject_classifier.utils.utils import xlsx_to_dataframe, dataframe_to_xlsx
from qsep_feedback_subject_classifier.row_collapser import collapse_rows
from qsep_feedback_subject_classifier import categorize_label
from utils.utils import unmerge_cells_and_fill, trim_rows_until_subject
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def categorize_subjects(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    Replace strings in the 'Subject' column with only the first word.
    
    Args:
        dataframe (pd.DataFrame): Input DataFrame with a 'Subject' column
        
    Returns:
        pd.DataFrame: DataFrame with modified 'Subject' column containing only first words
    """
    try:
        # Check if 'Subject' column exists
        if 'Subject' not in dataframe.columns:
            print("Warning: 'Subject' column not found in DataFrame")
            return dataframe
        
        # Create a copy to avoid modifying the original DataFrame
        modified_df = dataframe.copy()
        
        # Extract first word from each Subject entry
        subject_values = modified_df['Subject']
        first_words = []
        
        for subject_value in subject_values:
            # Convert to string and handle NaN/None values
            subject_string = str(subject_value)
            
            # Split by spaces and take the first word
            words = subject_string.split()
            if len(words) > 0:
                first_word = words[0]
            else:
                first_word = subject_string  # Keep original if no spaces found
            
            first_words.append(first_word)
        
        # Replace the Subject column with first words
        modified_df['Subject'] = first_words
        
        return modified_df
        
    except Exception as error:
        print(f"Error extracting first words from Subject column: {error}")
        return dataframe


def main():
    # Convert XLSM file to XLSX
    # input_file = r"docs\input\Katie_July_simple_titles.xlsm"
    input_file = pyperclip.paste().strip().strip('"')
    Path(input_file).expanduser().resolve()
    if not Path(input_file).exists():
        raise FileNotFoundError(f"File not found: {input_file}")
    wb = load_workbook(input_file)
    ws = wb.active  # Or use wb["SheetName"]

    unmerge_cells_and_fill(ws)
    trim_rows_until_subject(ws)
    wb.save("test_clean.xlsx")
    
    # output_file = r"docs\output\Katie_July_Collapsed.xlsx"
    
    # input_df = xlsx_to_dataframe(input_file)
    # input_df.insert(1, 'Categorized_Subject', input_df['Subject'].apply(categorize_label.categorize))
    # collapsed_df = collapse_rows(input_df)
    # output_df = dataframe_to_xlsx(collapsed_df, output_file)
    # print(f"Collapsed data saved to {output_file}")

if __name__ == "__main__":
    main()